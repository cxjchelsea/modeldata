# _*_ coding: utf-8 _*_
# @File:    dataku
# @Time:    2024/4/15 15:05
# @Author:  chenxuejiao
# @Contact: 15801102378@163.com
# @Version: V 0.1
import json
import requests
from flask import Flask, jsonify, send_file, Response, request
import mysql.connector
import math
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
import os
import sys
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

users = [
    {'username': 'admin', 'password': 'admin123', 'role': 'admin'},
    {'username': 'user', 'password': 'user123', 'role': 'user'},
]
key_mapping = {
    '模型名称（中文）': 'chineseName',
    '模型名称（英文）': 'englishName',
    '运行语言': 'language',
    '标签': 'label',
    '所属产线': 'productionLine',
    '所属工序': 'process',
    '参数描述': 'parameterDescription',
    '测试代码': 'predictCode',
    '测试结果样例': 'resultSample'
}

def connect_to_database():
    """连接到MySQL数据库"""
    conn = mysql.connector.connect(
        host='localhost',
        database='test',
        user='root',
        password='123456',
        auth_plugin='mysql_native_password'
    )
    return conn
def get_table_columns(table):
    try:
        # 连接到 MySQL 数据库
        conn = connect_to_database()

        if conn.is_connected():
            cursor = conn.cursor()

            # 执行 DESC 查询获取表结构信息
            query = f"DESCRIBE {table}"
            cursor.execute(query)

            # 读取查询结果
            columns = [column[0] for column in cursor.fetchall()]

            # 打印表头列名数组
            return columns

    except mysql.connector.Error as error:
        print("Error reading table columns from MySQL:", error)

    finally:
        cursor.close()
        conn.close()
def get_column_data(column, table):
    try:
        # 连接到 MySQL 数据库
        connection = connect_to_database()

        if connection.is_connected():
            cursor = connection.cursor()

            # 执行查询以获取特定列的数据
            query = f"SELECT {column} FROM {table}"
            cursor.execute(query)

            # 读取查询结果并将其存储在数组中
            column_data = [row[0] for row in cursor.fetchall()]

            return column_data

    except mysql.connector.Error as error:
        print("Error reading column data from MySQL:", error)
        return []

    finally:
        cursor.close()
        connection.close()
def get_row_count(table_name):
    try:
        # 连接到 MySQL 数据库
        connection = connect_to_database()

        if connection.is_connected():
            # 创建游标对象
            cursor = connection.cursor()

            # 执行查询
            cursor.execute(f"SELECT COUNT(*) FROM {table_name}")

            # 获取结果
            row_count = cursor.fetchone()[0]

            # 关闭游标和连接
            cursor.close()
            connection.close()

            return row_count
        else:
            print("连接失败！")
            return None

    except mysql.connector.Error as e:
        print(f"发生错误：{e}")
        return None
def get_column_name_by_index(table_name, column_index):
    try:
        # 连接到 MySQL 数据库
        connection = connect_to_database()

        if connection.is_connected():
            cursor = connection.cursor()

            # 执行 SQL 查询以获取表的列名和类型
            cursor.execute("SHOW COLUMNS FROM {}".format(table_name))

            # 获取所有列名和类型的元组列表
            columns_info = cursor.fetchall()

            # 检查索引是否有效
            if 0 <= column_index < len(columns_info):
                column_name = columns_info[column_index][0]

                # 关闭游标和连接
                cursor.close()
                connection.close()

                return column_name
            else:
                print("无效的列索引")
                return None
        else:
            print("无法连接到数据库")
            return None
    except Exception as e:
        print("发生错误:", e)
        return None
def check_production_line(df, column_name, production_lines):
    values = df[column_name].unique()
    for value in values:
        if value in production_lines:
            return True
    return False

def compare_headers(expected_headers, uploaded_headers):
    return expected_headers == uploaded_headers

def extract_from_excel(file_path):
    # 读取Excel文件
    df = pd.read_excel(file_path)

    nested_array = df.values.tolist()

    return nested_array

def insert_data_to_mysql(data, table):
    try:
        # 连接 MySQL 数据库
        conn = connect_to_database()
        cursor = conn.cursor()

        # 校验数据是否为空
        if not data:
            print("要插入的数据为空。")
            return

        # 校验数据格式并获取列数
        num_columns = len(data[0])
        for row in data:
            if len(row) != num_columns:
                print("数据行长度与目标表的列数不匹配。")
                return

        # 将空值替换为 None
        data_with_none = [[None if pd.isnull(value) else value for value in row] for row in data]

        # 插入数据到表中
        placeholders = ', '.join(['%s'] * num_columns)
        query = f"INSERT INTO {table} VALUES ({placeholders})"
        cursor.executemany(query, data_with_none)

        # 提交事务
        conn.commit()
        print("数据已成功插入到数据库表中。")

    except mysql.connector.Error as err:
        print("MySQL 错误:", err)
        conn.rollback()  # 回滚事务以撤销之前的操作

    finally:
        # 关闭连接
        cursor.close()
        conn.close()

def query_rows_by_ids(ids, table):
    # 建立数据库连接
    conn = connect_to_database()

    try:
        # 创建游标对象
        cursor = conn.cursor()

        # 构建 SQL 查询语句，并使用参数占位符来传递表名和 id 数组
        sql = f"SELECT * FROM {table} WHERE ID IN ({', '.join(['%s'] * len(ids))})"

        # 执行 SQL 语句，传递 id 数组作为参数
        cursor.execute(sql, tuple(ids))

        # 获取查询结果
        rows = cursor.fetchall()

        # 将查询结果转换为数组形式输出
        result = []
        for row in rows:
            result.append(list(row))

        return result

    except mysql.connector.Error as error:
        # 处理查询失败的情况
        print("查询数据行时出现错误:", error)

    finally:
        # 关闭游标和数据库连接
        if cursor:
            cursor.close()
        if conn.is_connected():
            conn.close()
def query_range(table, selected_id):
    conn = connect_to_database()
    cursor = conn.cursor()
    query_max_id = f"SELECT MAX(ID) FROM {table}"
    cursor.execute(query_max_id)
    max_id = int(cursor.fetchone()[0])  # 将最大ID转换为整数

    if int(selected_id) > max_id:  # 将selected_id转换为整数再进行比较
        query = f"SELECT * FROM {table} WHERE ID = %s"
        cursor.execute(query, (max_id,))
    else:
        query = f"SELECT * FROM {table} WHERE ID = %s"
        cursor.execute(query, (selected_id,))

    data = cursor.fetchall()
    cursor.close()
    conn.close()

    if int(selected_id) > max_id:  # 同样需要将selected_id转换为整数
        min_val = data[0][2]
        max_val = float('inf')
    else:
        min_val = data[0][1]
        max_val = data[0][2]

    return min_val, max_val
extract_from_excel
def query_data_in_range(table, column, min_value, max_value, page, page_size):
    # 连接到MySQL数据库
    conn = connect_to_database()
    # 创建游标对象
    cursor = conn.cursor()

    try:
        offset = (page - 1) * page_size
        if max_value == float('inf'):
            query = f"SELECT * FROM {table} WHERE {column} >= %s LIMIT %s OFFSET %s"
            cursor.execute(query, (min_value, page_size, offset))
        else:
            query = f"SELECT * FROM {table} WHERE {column} >= %s AND {column} < %s LIMIT %s OFFSET %s"
            cursor.execute(query, (min_value, max_value, page_size, offset))
        results = cursor.fetchall()

        if max_value == float('inf'):
            count_query = f"SELECT COUNT(*) FROM {table} WHERE {column} >= %s"
            cursor.execute(count_query, (min_value,))
        else:
            count_query = f"SELECT COUNT(*) FROM {table} WHERE {column} >= %s AND {column} < %s"
            cursor.execute(count_query, (min_value, max_value))
        total_rows = cursor.fetchone()[0]

        # 计算总页数
        total_pages = math.ceil(total_rows / page_size)

        return {'Data': results, 'totalPages': total_pages}

    except mysql.connector.Error as err:
        print("MySQL错误:", err)

    finally:
        # 关闭游标和连接
        cursor.close()
        conn.close()

def delete_rows_by_ids(ids,table):
    # 建立数据库连接
    conn = connect_to_database()

    try:
        # 创建游标对象
        cursor = conn.cursor()

        # 构建 SQL 删除语句，并使用参数占位符来传递表名和 id 数组
        sql = f"DELETE FROM {table} WHERE ID IN ({', '.join(['%s'] * len(ids))})"

        # 执行 SQL 语句，传递 id 数组作为参数
        cursor.execute(sql, tuple(ids))

        # 提交事务
        conn.commit()


    except mysql.connector.Error as error:
        # 处理删除失败的情况
        print("删除数据行时出现错误:", error)

    finally:
        # 关闭游标和数据库连接
        if cursor:
            cursor.close()
        if conn.is_connected():
            conn.close()

def edit_rows_by_ids(table, column, old_values, columns, new_values):
    # 连接 MySQL 数据库
    conn = connect_to_database()
    cursor = conn.cursor()
    try:
        # 循环处理每一行新数据
        for i, old_value in enumerate(old_values):
            # 构建 SQL 更新语句
            update_query = f"UPDATE {table} SET "
            for col in columns:
                update_query += f"`{col}` = %s, "  # 使用反引号括起列名
            update_query = update_query[:-2]  # 去除最后一个逗号和空格
            update_query += f" WHERE `{column}` = %s"  # 使用反引号括起列名

            # 将新行的值添加到参数列表中
            params = new_values[i] + [old_value]

            # 执行更新操作
            cursor.execute(update_query, params)

        # 提交事务
        conn.commit()

        print(f"成功更新 {len(old_values)} 行.")

    except mysql.connector.Error as err:
        print(f"更新失败: {err}")

    finally:
        cursor.close()
        conn.close()
def modify_table_headers(table, new_column_names):
    try:
        # 连接到 MySQL 数据库
        conn = connect_to_database()

        # 创建游标
        cursor = conn.cursor()

        # 获取当前表的列名和数据类型信息
        cursor.execute(f"SHOW COLUMNS FROM {table}")
        current_columns_info = cursor.fetchall()

        # 从列信息中提取列名和数据类型
        current_column_names = [column[0] for column in current_columns_info]

        # 检查输入数组和当前列名的长度是否一致
        if len(new_column_names) != len(current_column_names):
            print("输入数组和当前列名的长度不一致。")
            return

        # 查找并保留序号列
        index_column = None
        for i, column_info in enumerate(current_columns_info):
            if column_info[0] == '序号':
                index_column = current_column_names.pop(i)
                current_columns_info.pop(i)
                break

        # 构建修改列名的语句
        for i, (old_column_name, new_column_name) in enumerate(zip(current_column_names, new_column_names)):
            if old_column_name != new_column_name:
                alter_query = f"ALTER TABLE {table} CHANGE COLUMN {old_column_name} {new_column_name} {current_columns_info[i][1]}"
                cursor.execute(alter_query)

        # 添加序号列回表中
        if index_column:
            cursor.execute(f"ALTER TABLE {table} ADD COLUMN {index_column} INT FIRST")

        # 提交事务
        conn.commit()

        print("表头修改成功。")

    except mysql.connector.Error as err:
        print("Error:", err)

    finally:
        # 关闭游标和连接
        if 'cursor' in locals():
            cursor.close()
        if 'conn' in locals():
            conn.close()
def delete_column(table_name):
    conn = connect_to_database()
    cursor = conn.cursor()
    alter_query = f"ALTER TABLE {table_name} DROP COLUMN ID"
    cursor.execute(alter_query)
    conn.commit()
    cursor.close()
    conn.close()

def add_serial_number_column(table_name):
    conn = connect_to_database()
    cursor = conn.cursor()
    cursor.execute(f"SHOW COLUMNS FROM {table_name} LIKE 'ID'")
    result = cursor.fetchone()
    if result:
        cursor.execute(f"ALTER TABLE {table_name} DROP COLUMN ID")
    cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN ID INT AUTO_INCREMENT PRIMARY KEY FIRST")
    cursor.execute(f"SET @row_number = 0")
    cursor.execute(f"UPDATE {table_name} SET ID = (@row_number:=@row_number+1)")
    conn.commit()
    cursor.close()
    conn.close()
    print("序号列添加成功，并已生成序号信息。")

def update_ID(table_name):
    delete_column(table_name)
    add_serial_number_column(table_name)
def split_string(input_string):
    # 使用 split() 方法按照 '-' 进行分割
    parts = input_string.split('-')
    return parts
def count_rows_with_values(table_name, column1, value1, column2, value2):

    conn = connect_to_database()
    cursor = conn.cursor()

    if (value1 == '*' and value2 == '*'):
        # 执行查询
        query = f"SELECT COUNT(*) FROM {table_name}"
        cursor.execute(query)
    elif (value1 != '*' and value2 == '*'):
        # 执行查询
        query = f"SELECT COUNT(*) FROM {table_name} WHERE {column1} = '{value1}'"
        cursor.execute(query)
    else:
        # 执行查询
        query = f"SELECT COUNT(*) FROM {table_name} WHERE {column1} = '{value1}' AND {column2} = '{value2}'"
        cursor.execute(query)

    count = cursor.fetchone()[0]
    cursor.close()
    conn.close()
    return count
def count_rows_by_search(search_keywords, search_model):
    conn = connect_to_database()
    cursor = conn.cursor()

    try:
        # 构建查询条件
        conditions = []
        query_parameters = []

        for keyword in search_keywords:
            conditions.append("模型名称（中文） LIKE %s")
            query_parameters.append('%' + keyword + '%')

        # 如果有模型名称限制，则添加到条件中
        if search_model:
            conditions.append("标签 = %s")
            query_parameters.append(search_model)

        # 构建完整的查询语句
        query = f"SELECT COUNT(*) FROM modelist WHERE {' AND '.join(conditions)}"

        # 执行查询语句
        cursor.execute(query, query_parameters)

        # 获取查询结果
        count = cursor.fetchone()[0]

        return count

    except mysql.connector.Error as err:
        print("Error:", err)

    finally:
        # 关闭游标和连接
        cursor.close()
        conn.close()
def extract_data(column1, value1, column2, value2, columns_to_select, column_mapping):
    try:
        # 连接到MySQL数据库
        conn = connect_to_database()

        cursor = conn.cursor()
        if (value1 == '*' and value2 == '*'):
            # 执行查询
            query = f"SELECT {', '.join(columns_to_select)} FROM modelist"
            cursor.execute(query)
        elif (value1 != '*' and value2 == '*'):
            # 执行查询
            query = f"SELECT {', '.join(columns_to_select)} FROM modelist WHERE {column1} = '{value1}'"
            cursor.execute(query)
        else:
            # 执行查询
            query = f"SELECT {', '.join(columns_to_select)} FROM modelist WHERE {column1} = '{value1}' AND {column2} = '{value2}'"
            cursor.execute(query)


        # 处理结果并生成嵌套的键值对
        result = {}
        for row in cursor.fetchall():
            key = row[0]  # 第一列作为外部字典的键
            inner_dict = {}
            for i, column in enumerate(columns_to_select[1:], start=1):
                # 将列名映射为英文
                english_column_name = column_mapping.get(column, column)
                inner_dict[english_column_name] = row[i]  # 内部字典的键是映射后的英文列名，值是列值
            result[int(key)-1] = inner_dict

        return result
    except mysql.connector.Error as err:
        print(f"Error: {err}")
        raise
    finally:
        # 关闭连接
        if 'conn' in locals():
            conn.close()
def search_in_database(search_keywords, search_model, columns_to_select, column_mapping):
    # 连接数据库
    conn = connect_to_database()
    # 创建游标对象
    cursor = conn.cursor()

    try:
        # 构建查询条件
        conditions = []
        query_parameters = []

        for keyword in search_keywords:
            conditions.append("模型名称（中文） LIKE %s")
            query_parameters.append('%' + keyword + '%')

        # 如果有模型名称限制，则添加到条件中
        if search_model:
            conditions.append("标签 = %s")
            query_parameters.append(search_model)

        # 构建完整的查询语句
        query = f"SELECT {', '.join(columns_to_select)} FROM modelist WHERE {' AND '.join(conditions)}"

        # 执行查询语句
        cursor.execute(query, query_parameters)

        # 获取查询结果
        result = {}
        for row in cursor.fetchall():
            key = row[0]  # 第一列作为外部字典的键
            inner_dict = {}
            for i, column in enumerate(columns_to_select[1:], start=1):
                # 将列名映射为英文
                english_column_name = column_mapping.get(column, column)
                inner_dict[english_column_name] = row[i]  # 内部字典的键是映射后的英文列名，值是列值
            result[int(key) - 1] = inner_dict

        return result

    except mysql.connector.Error as err:
        print("Error:", err)

    finally:
        # 关闭游标和连接
        cursor.close()
        conn.close()
def fetch_data_from_mysql(column, value, select_columns, key_column):
    # 连接到MySQL数据库（用你的实现替换connect_to_database）
    conn = connect_to_database()
    cursor = conn.cursor()

    try:
        # 将select_columns转换为逗号分隔的字符串
        select_columns_str = ', '.join(select_columns)

        # 执行SQL查询
        query = f"SELECT {key_column}, {select_columns_str} FROM modelist WHERE {column} = %s"
        cursor.execute(query, (value,))

        # 检索查询结果
        result = cursor.fetchall()
        data_dict = {}

        if result:
            # 用键-值对填充字典，并将值存入数组
            for row in result:
                key_value = row[0]  # 假设第一列是键
                values = row[1:]
                data_dict[key_value] = dict(zip(select_columns, values))
                values_array = list(data_dict.values())
                values_dict = values_array[0]
                value_value = []
                for value_dict in values_array:
                    value_value.extend(value_dict.values())
        else:
            print("未找到匹配记录.")

        return select_columns, values_dict, value_value

    except mysql.connector.Error as e:
        print(f"错误: {e}")

def update_mysql_data(data_array, fixed_field):
    # 连接到 MySQL 数据库
    conn = connect_to_database()

    try:
        with conn.cursor() as cursor:
            # 遍历数据数组中的每一个键值对
            for item in data_array:
                # 构造 SQL 更新语句
                sql = f"UPDATE modelist SET "
                for key, value in item.items():
                    if key != fixed_field:
                        sql += f"{key} = '{value}', "
                sql = sql[:-2]  # 去掉末尾的逗号和空格
                sql += f" WHERE {fixed_field} = '{item[fixed_field]}'"

                # 执行 SQL 更新语句
                cursor.execute(sql)

        # 提交事务
        conn.commit()

    finally:
        # 关闭数据库连接
        conn.close()

def get_same_directory_path(filename):
    """获取脚本或可执行程序所在目录的文件路径"""
    if getattr(sys, 'frozen', False):  # 判断是否为可执行程序
        exe_dir = os.path.dirname(sys.executable)
    else:  # 如果是脚本则获取脚本所在目录
        exe_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(exe_dir, filename)

def get_parameter_information(table_name, model_name):
    try:
        # 连接到 MySQL 数据库
        connection = connect_to_database()

        if connection.is_connected():
            # 创建游标对象
            cursor = connection.cursor()

            # 执行查询
            query = f"SELECT `参数描述` FROM `{table_name}` WHERE `模型名称（中文）` = %s"
            cursor.execute(query, (model_name,))
            describe = cursor.fetchone()[0]
            query = f"SELECT `测试代码` FROM `{table_name}` WHERE `模型名称（中文）` = %s"
            cursor.execute(query, (model_name,))
            code = cursor.fetchone()[0]
            query = f"SELECT `参数类型` FROM `{table_name}` WHERE `模型名称（中文）` = %s"
            cursor.execute(query, (model_name,))
            type = cursor.fetchone()[0]
            # 关闭游标和连接
            cursor.close()
            connection.close()

            return describe, code, type
        else:
            print("连接失败！")
            return None

    except mysql.connector.Error as e:
        print(f"发生错误：{e}")
        return None
def create_tip(table_name, model_name, para_type, para_describe, key_value, key_dict):
    connection = connect_to_database()
    cursor = connection.cursor()
    query = f"SELECT `模型名称（英文）` FROM `{table_name}` WHERE `模型名称（中文）` = %s"
    cursor.execute(query, (model_name,))
    model_english = cursor.fetchone()[0]
    # 关闭游标和连接
    cursor.close()
    connection.close()
    template = f"""
# 动态库的导入
# 导入python的外部函数库ctypes提供与C兼容的数据类型以调用DLL。
# 将动态库dll文件的绝对路径赋给dll_path（直接替换path/to/your/dll）

import ctypes
dll_path = path/to/your/dll
model = ctypes.CDLL(dll_path)


# 模型函数的定义
# 根据函数自动生成参数定义。

model.{model_english}.argtypes = {para_type}
model.{model_english}.restype = {para_type[0]}


# 输入参数的赋值
# 此处根据需要给模型的输入参数赋值，替换模板中的输入变量的值。
{para_describe}

{key_value}


# 结果计算

result = model.{model_english}{key_dict}
# print(result)
    """
    return template

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()

    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({'error': 'Invalid username or password'}), 400

    user = next((user for user in users if user['username'] == username and user['password'] == password), None)

    if user:
        return jsonify({'message': 'Login successful', 'role': user['role']}), 200
    else:
        return jsonify({'error': 'Invalid username or password'}), 401

@app.route('/get_line_data', methods=['GET'])
def get_line_header():
    try:
        line_data_count = get_row_count("information_roll_line")

        conn = connect_to_database()
        cursor = conn.cursor()
        cursor.execute("SHOW COLUMNS FROM information_roll_line")
        global table_header
        table_header = [column[0] for column in cursor.fetchall()]
        cursor.close()

        cursor = conn.cursor()
        data_query = "SELECT * FROM information_roll_line"
        cursor.execute(data_query)
        table_data = cursor.fetchall()
        cursor.close()

        cursor = conn.cursor()
        query = "SELECT * FROM information_roll_line"  # 获取所有列的数据
        cursor.execute(query)
        data = cursor.fetchall()
        selected_line_index = 1
        selected_line_data = [row[selected_line_index] for row in data]
        selected_line_options = list(set(selected_line_data))
        cursor.close()

        cursor = conn.cursor()
        query = "SELECT * FROM information_roll_line"  # 获取所有列的数据
        cursor.execute(query)
        data = cursor.fetchall()
        selected_design_index = 2
        selected_design_data = [row[selected_design_index] for row in data]
        selected_design_options = list(set(selected_design_data))
        cursor.close()
        conn.close()
        return jsonify({'LineDataCount': line_data_count, 'tableHeader': table_header, 'tableData': table_data, 'selectedLineOptions': selected_line_options, 'selectedDesignOptions': selected_design_options})
    except Exception as e:
        print(f"Error fetching table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/send_selected_line', methods=['POST'])
def send_selected_line():
    try:
        data = request.get_json()
        selected_line = data.get('selectedLine')
        selected_line_index = 1
        column_name = get_column_name_by_index('information_roll_line', selected_line_index)
        conn = connect_to_database()
        cursor = conn.cursor()
        query = f"SELECT * FROM information_roll_line WHERE {column_name} = %s"
        cursor.execute(query, (selected_line,))
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)})
@app.route('/send_selected_design', methods=['POST'])
def send_selected_design():
    try:
        data = request.get_json()
        selected_design = data.get('selectedDesign')
        selected_design_index = 2  # 假设您想按照第三列（索引为2）的值来筛选
        column_name = get_column_name_by_index('information_roll_line', selected_design_index)
        conn = connect_to_database()
        cursor = conn.cursor()
        query = f"SELECT * FROM information_roll_line WHERE {column_name} = %s"  # 使用%s作为占位符
        cursor.execute(query, (selected_design,))
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)})
@app.route('/send_selected_year', methods=['POST'])
def send_selected_year():
    try:
        data = request.get_json()
        selected_year = data.get('selectedYear')
        selected_year_index = 3  # 假设您想按照第三列（索引为2）的值来筛选
        column_name = get_column_name_by_index('information_roll_line', selected_year_index)
        conn = connect_to_database()
        cursor = conn.cursor()
        query = f"SELECT * FROM information_roll_line WHERE {column_name} = %s"  # 使用%s作为占位符
        cursor.execute(query, (selected_year,))
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/upload_line', methods=['POST'])
def upload_line():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'})

    files = request.files.getlist('file')

    if not files:
        return jsonify({'error': '上传文件为空'})

    expected_headers = get_table_columns('information_roll_line')

    uploaded_files = []  # 用于存储上传成功的文件名
    mismatched_files = []  # 用于存储表头不一致的文件名
    existing_production_files = []  # 用于存储已存在于产线名称数组中的文件名
    production_lines = get_column_data('产线名称', 'information_roll_line')

    # 连接 MySQL 数据库
    conn = connect_to_database()
    cursor = conn.cursor()

    for file in files:
        filename = file.filename

        # 使用 pandas 读取 Excel 文件并获取表头
        df = pd.read_excel(file)
        uploaded_headers = df.columns.tolist()

        if compare_headers(expected_headers, uploaded_headers):
            # 检查是否存在于产线名称数组中
            if check_production_line(df, '产线名称', production_lines):
                existing_production_files.append(filename)
            else:
                table_data = extract_from_excel(file)

                insert_data_to_mysql(table_data, 'information_roll_line')
                uploaded_files.append(filename)
                update_ID('information_roll_line')
                # line_data_count = get_row_count("information_roll_line")
        else:
            mismatched_files.append(filename)

    data_query = "SELECT * FROM information_roll_line"
    cursor.execute(data_query)
    table_data = cursor.fetchall()
    # 关闭 MySQL 连接
    cursor.close()
    conn.close()

    return jsonify({'beingUpload': uploaded_files, 'mismatchUpload': mismatched_files,
                    'existUpload': existing_production_files,'tableData': table_data})
                    # ,"LineDataCount": line_data_count})

@app.route('/select_lines', methods=['POST'])
def select_lines():
    try:
        data = request.get_json()
        selected_rows_ids = data.get('selectedRowsIds')
        selected_data = query_rows_by_ids(selected_rows_ids, 'information_roll_line')
        return jsonify({'selectedData': selected_data})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/delete_lines', methods=['POST'])
def delete_lines():
    try:
        data = request.get_json()
        selected_rows_ids = data.get('selectedRowsIds')
        delete_rows_by_ids(selected_rows_ids, 'information_roll_line')
        update_ID('information_roll_line')
        conn = connect_to_database()
        cursor = conn.cursor()
        data_query = "SELECT * FROM information_roll_line"
        cursor.execute(data_query)
        table_data = cursor.fetchall()
        cursor.close()
        conn.close()
        line_data_count = get_row_count("information_roll_line")
        return jsonify({'tableData': table_data, 'LineDataCount': line_data_count})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/edit_lines', methods=['POST'])
def edit_lines():
    try:
        data = request.get_json()
        selected_rows_ids = data.get('selectedRowsIds')
        update_data = data.get('updateData')
        edit_rows_by_ids('information_roll_line', 'ID', selected_rows_ids, table_header, update_data)
        conn = connect_to_database()
        cursor = conn.cursor()
        data_query = "SELECT * FROM information_roll_line"
        cursor.execute(data_query)
        table_data = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({'tableData': table_data})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500
@app.route('/edit_line_header', methods=['POST'])
def edit_line_header():
    try:
        data = request.get_json()
        update_header = data.get('updateHeader')
        modify_table_headers('information_roll_line', update_header)
        conn = connect_to_database()
        cursor = conn.cursor()
        cursor.execute("SHOW COLUMNS FROM information_roll_line")
        table_header = [column[0] for column in cursor.fetchall()]
        cursor.close()
        conn.close()
        return jsonify({'tableHeader': table_header})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/export_line')
def export_line():
    try:
        conn = connect_to_database()
        cursor = conn.cursor()
        table_name = 'information_roll_line'
        cursor.execute(f"SELECT * FROM {table_name}")
        data = cursor.fetchall()
        wb = Workbook()
        ws = wb.active
        ws.append(cursor.column_names)  # 表头
        for row in data:
            ws.append(row)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={table_name}.xlsx"}
        )
    except Exception as e:
        return str(e)
    finally:
        if 'conn' in locals():
            conn.close()
###
@app.route('/get_material_data', methods=['GET'])
def get_material_header():
    try:
        material_data_count = get_row_count("material_data")

        conn = connect_to_database()
        cursor = conn.cursor()
        cursor.execute("SHOW COLUMNS FROM material_data")
        global table_header
        table_header = [column[0] for column in cursor.fetchall()]
        cursor.close()

        cursor = conn.cursor()
        data_query = "SELECT * FROM material_data"
        cursor.execute(data_query)
        table_data = cursor.fetchall()
        cursor.close()

        cursor = conn.cursor()
        query = "SELECT * FROM material_data"  # 获取所有列的数据
        cursor.execute(query)
        data = cursor.fetchall()
        selected_type_index = 3
        selected_type_data = [row[selected_type_index] for row in data]
        selected_type_options = list(set(selected_type_data))
        cursor.close()


        cursor = conn.cursor()
        query = "SELECT * FROM material_data"  # 获取所有列的数据
        cursor.execute(query)
        data = cursor.fetchall()
        selected_type1_index = 1
        selected_type1_data = [row[selected_type1_index] for row in data]
        selected_type1_options = list(set(selected_type1_data))
        cursor.close()

        cursor = conn.cursor()
        query = "SELECT * FROM material_data"  # 获取所有列的数据
        cursor.execute(query)
        data = cursor.fetchall()
        selected_type2_index = 2
        selected_type2_data = [row[selected_type2_index] for row in data]
        selected_type2_options = list(set(selected_type2_data))
        cursor.close()

        cursor = conn.cursor()
        query = "SELECT * FROM material_data"  # 获取所有列的数据
        cursor.execute(query)
        data = cursor.fetchall()
        selected_grade_index = 4
        selected_grade_data = [row[selected_grade_index] for row in data]
        selected_grade_options = list(set(selected_grade_data))
        cursor.close()


        conn.close()
        return jsonify({'MaterialDataCount': material_data_count, 'tableHeader': table_header, 'tableData': table_data, 'selectedTypeOptions': selected_type_options, 'selectedType1Options': selected_type1_options, 'selectedType2Options': selected_type2_options, 'selectedGradeOptions': selected_grade_options})
    except Exception as e:
        print(f"Error fetching table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/send_selected_type', methods=['POST'])
def send_selected_type():
    try:
        data = request.get_json()
        selected_type = data.get('selectedType')
        selected_type_index = 3  # 假设您想按照第三列（索引为2）的值来筛选
        column_name = get_column_name_by_index('material_data', selected_type_index)
        conn = connect_to_database()
        cursor = conn.cursor()
        query = f"SELECT * FROM material_data WHERE {column_name} = %s"  # 使用%s作为占位符
        cursor.execute(query, (selected_type,))
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/send_selected_type1', methods=['POST'])
def send_selected_type1():
    try:
        data = request.get_json()
        selected_type1 = data.get('selectedType1')
        selected_type1_index = 1  # 假设您想按照第三列（索引为2）的值来筛选
        column_name = get_column_name_by_index('material_data', selected_type1_index)
        conn = connect_to_database()
        cursor = conn.cursor()
        query = f"SELECT * FROM material_data WHERE {column_name} = %s"  # 使用%s作为占位符
        cursor.execute(query, (selected_type1,))
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/send_selected_type2', methods=['POST'])
def send_selected_type2():
    try:
        data = request.get_json()
        selected_type2 = data.get('selectedType2')
        selected_type2_index = 2  # 假设您想按照第三列（索引为2）的值来筛选
        column_name = get_column_name_by_index('material_data', selected_type2_index)
        conn = connect_to_database()
        cursor = conn.cursor()
        query = f"SELECT * FROM material_data WHERE {column_name} = %s"  # 使用%s作为占位符
        cursor.execute(query, (selected_type2,))
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/send_selected_grade', methods=['POST'])
def send_selected_grade():
    try:
        data = request.get_json()
        selected_grade = data.get('selectedGrade')
        selected_grade_index = 4  # 假设您想按照第三列（索引为2）的值来筛选
        column_name = get_column_name_by_index('material_data', selected_grade_index)
        conn = connect_to_database()
        cursor = conn.cursor()
        query = f"SELECT * FROM material_data WHERE {column_name} = %s"  # 使用%s作为占位符
        cursor.execute(query, (selected_grade,))
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/upload_material', methods=['POST'])
def upload_material():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'})

    files = request.files.getlist('file')

    if not files:
        return jsonify({'error': '上传文件为空'})

    expected_headers = get_table_columns('material_data')

    uploaded_files = []  # 用于存储上传成功的文件名
    mismatched_files = []  # 用于存储表头不一致的文件名
    existing_production_files = []  # 用于存储已存在于产线名称数组中的文件名
    material_grade = get_column_data('Grade', 'material_data')

    # 连接 MySQL 数据库
    conn = connect_to_database()
    cursor = conn.cursor()

    for file in files:
        filename = file.filename

        # 使用 pandas 读取 Excel 文件并获取表头
        df = pd.read_excel(file)
        uploaded_headers = df.columns.tolist()

        if compare_headers(expected_headers, uploaded_headers):
            # 检查是否存在于产线名称数组中
            if check_production_line(df, 'Grade', material_grade):
                existing_production_files.append(filename)
            else:
                table_data = extract_from_excel(file)
                insert_data_to_mysql(table_data, 'material_data')
                uploaded_files.append(filename)
                update_ID('material_data')
                material_data_count = get_row_count("material_data")
        else:
            mismatched_files.append(filename)

    data_query = "SELECT * FROM material_data"
    cursor.execute(data_query)
    table_data = cursor.fetchall()
    # 关闭 MySQL 连接
    cursor.close()
    conn.close()

    return jsonify({'beingUpload': uploaded_files, 'mismatchUpload': mismatched_files,
                    'existUpload': existing_production_files,'tableData': table_data, "MaterialDataCount": material_data_count})

@app.route('/select_materials', methods=['POST'])
def select_materials():
    try:
        data = request.get_json()
        selected_rows_ids = data.get('selectedRowsIds')
        selected_data = query_rows_by_ids(selected_rows_ids, 'material_data')
        return jsonify({'selectedData': selected_data})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/delete_materials', methods=['POST'])
def delete_materials():
    try:
        data = request.get_json()
        selected_rows_ids = data.get('selectedRowsIds')
        delete_rows_by_ids(selected_rows_ids, 'material_data')
        update_ID('material_data')
        conn = connect_to_database()
        cursor = conn.cursor()
        data_query = "SELECT * FROM material_data"
        cursor.execute(data_query)
        table_data = cursor.fetchall()
        cursor.close()
        conn.close()
        material_data_count = get_row_count("material_data")
        return jsonify({'tableData': table_data, 'MaterialDataCount': material_data_count})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/edit_materials', methods=['POST'])
def edit_materials():
    try:
        data = request.get_json()
        selected_rows_ids = data.get('selectedRowsIds')
        update_data = data.get('updateData')
        edit_rows_by_ids('material_data', 'ID', selected_rows_ids, table_header, update_data)
        conn = connect_to_database()
        cursor = conn.cursor()
        data_query = "SELECT * FROM material_data"
        cursor.execute(data_query)
        table_data = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify({'tableData': table_data})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500
@app.route('/edit_material_header', methods=['POST'])
def edit_material_header():
    try:
        data = request.get_json()
        update_header = data.get('updateHeader')
        modify_table_headers('material_data', update_header)
        conn = connect_to_database()
        cursor = conn.cursor()
        cursor.execute("SHOW COLUMNS FROM material_data")
        table_header = [column[0] for column in cursor.fetchall()]
        cursor.close()
        conn.close()
        return jsonify({'tableHeader': table_header})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/export_material')
def export_material():
    try:
        conn = connect_to_database()
        cursor = conn.cursor()
        table_name = 'material_data'
        cursor.execute(f"SELECT * FROM {table_name}")
        data = cursor.fetchall()
        wb = Workbook()
        ws = wb.active
        ws.append(cursor.column_names)  # 表头
        for row in data:
            ws.append(row)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={table_name}.xlsx"}
        )
    except Exception as e:
        return str(e)
    finally:
        if 'conn' in locals():
            conn.close()
###
@app.route('/get_product_data', methods=['GET'])
def get_product_data():
    try:
        product_line_count = get_row_count('product_roll_line_name')
        conn = connect_to_database()
        cursor = conn.cursor()
        query = "SELECT roll_line, line_number FROM product_roll_line_name"
        cursor.execute(query)
        product_line_data = {row[0]: row[1] for row in cursor.fetchall()}
        total_product_line_data = sum(int(value) for value in product_line_data.values())

        roll_line_query = f"SELECT * FROM product_roll_line_name"
        cursor.execute(roll_line_query)
        roll_line = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return jsonify({'ProductLineCount': product_line_count, 'TotalProductLineData': total_product_line_data, 'ProductLineData': product_line_data, 'selectedPlineOptions': roll_line})
    except Exception as e:
        print(f"Error fetching table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/send_selected_pline', methods=['POST'])
def send_selected_pline():
    try:
        data = request.get_json()
        global selected_pline, tgt_wid, tgt_thick
        selected_pline = data.get('selectedPline')

        if selected_pline == 'jt2250':
            tgt_wid = 'T_Wid'
            tgt_thick = 'T_Thick'
        elif selected_pline == 'jinma1450':
            tgt_wid = 'fmx_tgt_width'
            tgt_thick = 'fmx_tgt_thick'
        elif selected_pline == 'qiangang1580':
            tgt_wid = 'pdo目标宽度H21_0'
            tgt_thick = 'pdo目标厚度H21_0'
        elif selected_pline == 'qiangang2260':
            tgt_wid = 'pdo目标宽度H11_0'
            tgt_thick = 'pdo目标厚度H11_0'
        elif selected_pline == 'shengyang1700':
            tgt_wid = '目标宽度'
            tgt_thick = '目标厚度'
        elif selected_pline == 'shagang':
            tgt_wid = 'PRO_WIDTH'
            tgt_thick = 'PRO_THICKNESS'

        select_product_line_count = get_row_count(selected_pline)
        page = data.get('page')
        page_size = data.get('pageSize')
        offset = (page - 1) * page_size
        conn = connect_to_database()
        cursor = conn.cursor()
        cursor.execute(f"SHOW COLUMNS FROM {selected_pline}")
        global table_header
        table_header = [column[0] for column in cursor.fetchall()]
        count_query = f"SELECT COUNT(*) FROM {selected_pline}"
        cursor.execute(count_query)
        total_rows = cursor.fetchone()[0]
        data_query = f"SELECT * FROM {selected_pline} LIMIT %s OFFSET %s"
        cursor.execute(data_query, (page_size, offset))
        table_data = cursor.fetchall()
        grade_query = f"SELECT Grade_name FROM {selected_pline}"
        cursor.execute(grade_query)
        data = [row[0] for row in cursor.fetchall() if row[0] is not None]
        selected_grade_name_options = list(sorted(set(data)))
        width_query = f"SELECT * FROM width"
        cursor.execute(width_query)
        selected_width_options = cursor.fetchall()
        max_width_option = selected_width_options[-1][2]
        thick_query = f"SELECT * FROM thick"
        cursor.execute(thick_query)
        selected_thick_options = cursor.fetchall()
        max_thick_option = selected_thick_options[-1][2]
        cursor.close()
        conn.close()
        return jsonify({'SelectProductLineCount': select_product_line_count, 'tableHeader': table_header, 'tableData': table_data, 'totalRows': total_rows, 'selectedGradenameOptions': selected_grade_name_options, 'selectedWidthOptions': selected_width_options, 'selectedThickOptions': selected_thick_options, 'maxWidthOption': max_width_option, 'maxThickOption': max_thick_option})
    except Exception as e:
        return jsonify({'error': str(e)})


@app.route('/send_selected_grade_name', methods=['POST'])
def send_selected_grade_name():
    try:
        data = request.get_json()
        selected_grade_name = data.get('selectedGradename')
        page = data.get('page')
        page_size = data.get('pageSize')
        conn = connect_to_database()
        cursor = conn.cursor()
        offset = (page - 1) * page_size
        query = f"SELECT * FROM {selected_pline} WHERE Grade_name = %s LIMIT %s OFFSET %s"
        cursor.execute(query, (selected_grade_name, page_size, offset))
        data = cursor.fetchall()
        cursor.close()
        conn.close()
        conn = connect_to_database()
        cursor = conn.cursor()
        count_query = f"SELECT COUNT(*) FROM {selected_pline} WHERE Grade_name = %s"
        cursor.execute(count_query, (selected_grade_name,))
        total_rows = cursor.fetchone()[0]
        cursor.close()
        conn.close()
        total_pages = math.ceil(total_rows / page_size)

        return jsonify({'Data': data, 'totalPages': total_pages})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/send_selected_width', methods=['POST'])
def send_selected_width():
    try:
        data = request.get_json()
        selected_width = data.get('selectedWidth')
        min_width, max_width = query_range('width', selected_width)
        page = data.get('page')
        page_size = data.get('pageSize')
        result = query_data_in_range(selected_pline, tgt_wid, min_width, max_width, page, page_size)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)})
@app.route('/send_selected_thick', methods=['POST'])
def send_selected_thick():
    try:
        data = request.get_json()
        selected_thick = data.get('selectedThick')
        min_thick, max_thick = query_range('thick', selected_thick)
        page = data.get('page')
        page_size = data.get('pageSize')
        result = query_data_in_range(selected_pline, tgt_thick, min_thick, max_thick, page, page_size)
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/upload_product', methods=['POST'])
def upload_product():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'})

    files = request.files.getlist('file')

    if not files:
        return jsonify({'error': '上传文件为空'})

    expected_headers = get_table_columns(selected_pline)

    uploaded_files = []  # 用于存储上传成功的文件名
    mismatched_files = []  # 用于存储表头不一致的文件名
    existing_production_files = []  # 用于存储已存在于产线名称数组中的文件名
    production_lines = get_column_data('Coil_ID', selected_pline)

    # 连接 MySQL 数据库
    conn = connect_to_database()
    cursor = conn.cursor()

    for file in files:
        filename = file.filename

        # 使用 pandas 读取 Excel 文件并获取表头
        df = pd.read_excel(file)
        uploaded_headers = df.columns.tolist()

        if compare_headers(expected_headers, uploaded_headers):
            # 检查是否存在于产线名称数组中
            if check_production_line(df, 'Coil_ID', production_lines):
                existing_production_files.append(filename)
            else:
                table_data = extract_from_excel(file)
                insert_data_to_mysql(table_data, selected_pline)
                uploaded_files.append(filename)
                update_ID(selected_pline)
                select_product_line_count = get_row_count(f"{selected_pline}")
        else:
            mismatched_files.append(filename)

    data_query = f"SELECT * FROM {selected_pline}"
    cursor.execute(data_query)
    table_data = cursor.fetchall()
    # 关闭 MySQL 连接
    cursor.close()
    conn.close()

    return jsonify({'beingUpload': uploaded_files, 'mismatchUpload': mismatched_files,
                    'existUpload': existing_production_files,'tableData': table_data, "SelectProductLineCount": select_product_line_count})

@app.route('/select_products', methods=['POST'])
def select_products():
    try:
        data = request.get_json()
        selected_rows_ids = data.get('selectedRowsIds')
        selected_data = query_rows_by_ids(selected_rows_ids, selected_pline)
        return jsonify({'selectedData': selected_data})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/delete_products', methods=['POST'])
def delete_products():
    try:
        data = request.get_json()
        selected_rows_ids = data.get('selectedRowsIds')
        delete_rows_by_ids(selected_rows_ids, selected_pline)
        update_ID(selected_pline)
        conn = connect_to_database()
        cursor = conn.cursor()
        data_query = f"SELECT * FROM {selected_pline}"
        cursor.execute(data_query)
        table_data = cursor.fetchall()
        cursor.close()
        conn.close()
        product_data_count = get_row_count(f"{selected_pline}")

        return jsonify({'tableData': table_data, "ProductDataCount": product_data_count})
    except Exception as e:
        print(f"Error delete table data: {e}")
        return jsonify({'error': 'Internal Server Error'}), 500

@app.route('/export_product')
def export_product():
    try:
        conn = connect_to_database()
        cursor = conn.cursor()
        export_gradename = request.args.get('exportGradename', '')
        export_width = request.args.get('exportWidth', '')
        export_thick = request.args.get('exportThick', '')
        table_name = selected_pline
        sql = f"SELECT * FROM {table_name} WHERE 1=1"
        if export_gradename:
            sql += f" AND Grade_name = '{export_gradename}'"
        if export_width:
            min_width, max_width = query_range('width', export_width)
            sql += f" AND {tgt_wid} >= {min_width} AND {tgt_wid} < {max_width}"
        if export_thick:
            min_thick, max_thick = query_range('thick', export_thick)
            sql += f" AND {tgt_thick} >= {min_thick} AND {tgt_wid} < {max_thick}"
        cursor.execute(sql)
        data = cursor.fetchall()
        wb = Workbook()
        ws = wb.active
        ws.append(cursor.column_names)
        for row in data:
            ws.append(row)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={table_name}.xlsx"}
        )
    except Exception as e:
        return str(e)
    finally:
        # 关闭数据库连接
        if 'conn' in locals():
            conn.close()


###
@app.route('/send_submenu', methods=['POST'])
def send_submenu():
    try:
        data = request.get_json()
        global current_page, local_current_pageNumber
        current_page = data.get('currentPage')
        local_current_pageNumber = data.get('localCurrentPageNumber')
        return jsonify({'currentPage': current_page, 'localCurrentPageNumber': local_current_pageNumber})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/send_search', methods=['POST'])
def send_search():
    try:
        data = request.get_json()
        request_body = data.get('requestBody')
        global search_keywords, search_model
        keys_present = [key for key in ['searchKeywords', 'model'] if key in request_body]
        if not keys_present:
            return jsonify({"message": "Missing searchKeywords"})
        elif len(keys_present) == 1:
            key = keys_present[0]
            if key == 'searchKeywords':
                search_keywords = request_body['searchKeywords']
                search_keywords = search_keywords.split(",")
                search_model = ''
                return jsonify({'searchKeywords': search_keywords, 'searchModel': search_model})
            elif key == 'model':
                return jsonify({"message": "Missing searchKeyword"})
        else:
            search_keywords = request_body['searchKeywords']
            search_keywords = search_keywords.split(",")
            search_model = request_body['model']
            return jsonify({'searchKeywords': search_keywords, 'searchModel': search_model})
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/send_items_per_page', methods=['POST'])
def send_items_per_page():
    try:
        data = request.get_json()
        items_per_page = data.get('itemsPerPage')
        # 根据当前页和每页项数计算起始和结束索引
        start_index = (local_current_pageNumber - 1) * items_per_page
        end_index = start_index + items_per_page
        if current_page == '模型':
            result = extract_data('标签', '*', '所属产线', '*', ['序号','模型名称（中文）', '模型名称（英文）', '运行语言', '标签', '所属产线'], key_mapping)
            result_list = list(result.values())
            paginated_result = result_list[start_index:end_index]
            current_model_num = count_rows_with_values('modelist', '标签', '*', '所属产线', '*')
            total_pages = current_model_num // 16 + 1
            model_num = get_row_count('modelist')
            if paginated_result is not None and paginated_result:
                print(f"'标签' 列与 '{current_page}' 相同的数据行的选择列：")
                print(paginated_result)
            else:
                print(f"'标签' 列与 '{current_page}' 相同的数据行不存在或选择失败.")
            return jsonify({'data': paginated_result, 'totalPages': total_pages, 'currentModelnum': current_model_num, 'modelNum': model_num})
        elif '-' not in current_page:
            label = split_string(current_page)[0]
            result = extract_data('标签', label, '所属产线', '*', ['序号','模型名称（中文）', '模型名称（英文）', '运行语言', '标签', '所属产线'], key_mapping)
            result_list = list(result.values())
            paginated_result = result_list[start_index:end_index]
            current_model_num = count_rows_with_values('modelist', '标签', label, '所属产线', '*')
            total_pages = current_model_num // 16 + 1
            model_num = get_row_count('modelist')
            if paginated_result is not None and paginated_result:
                print(f"'标签' 列与 '{current_page}' 相同的数据行的选择列：")
                print(paginated_result)
            else:
                print(f"'标签' 列与 '{current_page}' 相同的数据行不存在或选择失败.")
            return jsonify({'data': paginated_result, 'totalPages': total_pages, 'currentModelnum': current_model_num, 'modelNum': model_num})
        else:
            label = split_string(current_page)[0]
            pline = split_string(current_page)[1]
            result = extract_data('标签', label, '所属产线', pline, ['序号','模型名称（中文）', '模型名称（英文）', '运行语言', '标签', '所属产线'], key_mapping)
            result_list = list(result.values())
            paginated_result = result_list[start_index:end_index]
            current_model_num = count_rows_with_values('modelist', '标签', label, '所属产线', pline)
            total_pages = current_model_num // 16 + 1
            model_num = get_row_count('modelist')
            if paginated_result is not None and paginated_result:
                print(f"'标签' 列与 '{current_page}' 相同的数据行的选择列：")
                print(paginated_result)
            else:
                print(f"'标签' 列与 '{current_page}' 相同的数据行不存在或选择失败.")
            return jsonify({'data': paginated_result, 'totalPages': total_pages, 'currentModelnum': current_model_num, 'modelNum': model_num})

    except Exception as e:
        return jsonify({'error': str(e)})
@app.route('/send_search_items', methods=['POST'])
def send_search_items():
    try:
        data = request.get_json()
        items_per_page = data.get('itemsPerPage')
        # 根据当前页和每页项数计算起始和结束索引
        start_index = (local_current_pageNumber - 1) * items_per_page
        end_index = start_index + items_per_page
        result = search_in_database(search_keywords, search_model, ['序号', '模型名称（中文）', '模型名称（英文）', '运行语言', '标签', '所属产线'], key_mapping)
        result_list = list(result.values())
        paginated_result = result_list[start_index:end_index]
        current_model_num = count_rows_by_search(search_keywords, search_model)
        total_pages = current_model_num // 16 + 1
        model_num = get_row_count('modelist')
        if paginated_result is not None and paginated_result:
            print(f"'模型名称' 列与 '{search_keywords}' 相同的数据行的选择列：")
            print(paginated_result)
        else:
            print(f"'模型名称' 列与 '{search_keywords}' 相同的数据行不存在或选择失败.")
        return jsonify({'data': paginated_result, 'totalPages': total_pages, 'currentModelnum': current_model_num, 'modelNum': model_num})

    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/getTableData', methods=['POST'])
def get_table_data():
    data = request.get_json()
    global chinese_name
    chinese_name = data.get('chineseName')
    table_one = fetch_data_from_mysql('模型名称（中文）', chinese_name, ['模型名称（英文）', '所属产线'], '序号')[1]
    table_two = fetch_data_from_mysql('模型名称（中文）', chinese_name, ['运行语言', '所属工序'], '序号')[1]
    table_three = fetch_data_from_mysql('模型名称（中文）', chinese_name, ['参数描述'], '序号')[1]
    table_four = fetch_data_from_mysql('模型名称（中文）', chinese_name, ['测试代码'], '序号')[1]
    inbound_time = fetch_data_from_mysql('模型名称（中文）', chinese_name, ['入库时间'], '序号')[1]['入库时间']
    test_code = fetch_data_from_mysql('模型名称（中文）', chinese_name, ['测试代码'], '序号')[1]['测试代码']
    return jsonify({'tableOne': table_one, 'tableTwo': table_two, 'tableThree': table_three, 'tableFour': table_four,
                    'inboundTime': inbound_time, 'testCode': test_code})
@app.route('/get_pdf', methods=['GET'])
def get_pdf():

    executable_dir = get_same_directory_path('pdf')
    pdf_file_path = os.path.join(executable_dir, f"{chinese_name}.pdf")
    return send_file(pdf_file_path, as_attachment=False)

@app.route('/generate_code', methods=['GET'])
def generate_code():
    para_describe, code, para_type = get_parameter_information('modelist', chinese_name)
    array = [type.strip() for type in para_type.split(',')]
    para_type = [f"ctypes.c_{type}" for type in array]
    para = json.loads(code)
    req_data = para["req_data"]
    key_value = ", ".join(f"{key}={value}" for key, value in req_data.items())
    key_dict = f"({', '.join(req_data.keys())})"
    template = create_tip('modelist', chinese_name, para_type, para_describe, key_value, key_dict)
    print(template)
    return jsonify(template)
@app.route('/predict', methods=['POST'])
def predict():
    data = request.get_json()
    req_data = data.get('inputData')
    data = json.loads(req_data)
    url = fetch_data_from_mysql('模型名称（中文）', chinese_name, ['API'], '序号')[1]['API']
    headers = {"Content-Type": "application/json;charset=UTF-8"}
    result = requests.post(url=url, data=json.dumps(data), headers=headers, verify=False)
    resp_data = json.loads(result.text)["resp_data"]
    resp_data = list(resp_data.values())
    return jsonify(resp_data)
@app.route('/send_list', methods=['POST'])
def send_list():
    try:
        data = request.get_json()
        selected_label = data.get('selectedLabel')
        model_list = fetch_data_from_mysql('标签', selected_label, ['模型名称（中文）'], '序号')[2]
        return jsonify({'modelList': model_list})
    except Exception as e:
        return jsonify({'error': str(e)})
@app.route('/get_model_data', methods=['POST'])
def get_model_data():
    data = request.get_json()
    selected_model = data.get('selectedModel')
    models_data = []
    key = fetch_data_from_mysql('模型名称（中文）', selected_model[0], ['模型名称（中文）', '所属产线', '所属工序', '入库时间'], '序号')[0]
    for model in selected_model:
        model_data = fetch_data_from_mysql('模型名称（中文）', model, ['模型名称（中文）', '所属产线', '所属工序', '入库时间'], '序号')[1]
        models_data.append(model_data)
    return jsonify({'modelData': models_data, 'Key': key})
@app.route('/send_edit', methods=['POST'])
def send_edit():
    try:
        data = request.get_json()
        model_data = data.get('modelData')
        update_mysql_data(model_data, '模型名称（中文）')
        return jsonify({'message': 'success'})
    except Exception as e:
        return jsonify({'error': str(e)})
@app.route('/get_dll', methods=['GET'])
def get_dll():
    executable_dir = get_same_directory_path('dll')
    connection = connect_to_database()
    cursor = connection.cursor()
    query = f"SELECT `模型名称（英文）` FROM modelist WHERE `模型名称（中文）` = %s"
    cursor.execute(query, (chinese_name,))
    english_name = cursor.fetchone()[0]
    cursor.close()
    connection.close()
    dll_file_path = os.path.join(executable_dir, f"{english_name}.dll")
    return send_file(dll_file_path, as_attachment=False)

if __name__ == '__main__':
    # app.run(
    #     host='192.168.151.199',
    #     port=9999,
    #     debug=True
    # )
    app.run(
        host='0.0.0.0',
        port=60000,
        debug=True
    )